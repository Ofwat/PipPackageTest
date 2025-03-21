"""
Panacea code

Function used to do initial validation of the Excel files
"""
import uuid
from typing import Dict, Any, List
import logging
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import pandas as pd

# Configure logging for the function
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def validate_tabs_between_spreadsheets(spreadsheet1, spreadsheet2):
    """
    Compares the sheet names between two openpyxl workbook objects to check if they are identical.

    This function compares sheet names in both workbooks, ensuring that they contain the same tabs
    (ignoring order). If there are any missing tabs in either workbook,
        it will return False and provide
    details on which sheets are missing from each spreadsheet.

    Args:
        spreadsheet1 (openpyxl.workbook.workbook.Workbook): The first workbook object to compare.
        spreadsheet2 (openpyxl.workbook.workbook.Workbook): The second workbook object to compare.

    Returns:
        dict:
            - "status": "Ok" if both workbooks have the same sheet names, "Error" otherwise.
            - "description": A general description of the comparison result.
            - "errors": A dictionary containing detailed error messages about missing tabs.
              If no errors, this will be an empty dictionary.

    Raises:
        ValueError: If either argument is not a valid openpyxl workbook object.
        InvalidFileException: If there is an issue with loading the workbook.
        Exception: For any unexpected errors during execution.
    """
    # Validate input types
    if not isinstance(spreadsheet1, Workbook) or not isinstance(spreadsheet2, Workbook):
        raise ValueError("Both arguments must be valid openpyxl workbook objects.")

    # Get sheet names from both workbooks
    sheets1 = set(spreadsheet1.sheetnames)
    sheets2 = set(spreadsheet2.sheetnames)

    # Check for missing sheets in both spreadsheets
    missing_in_1 = sheets2 - sheets1
    missing_in_2 = sheets1 - sheets2

    result = {
        "status": "Ok",
        "description": "Both spreadsheets have the same sheet names.",
        "errors": {}
    }

    if missing_in_1 or missing_in_2:
        result["status"] = "Error"
        result["description"] = "Spreadsheets have different sheet names."
        errors = {}
        if missing_in_1:
            errors["Missing In Spreadsheet 1"] = list(missing_in_1)
        if missing_in_2:
            errors["Missing In Spreadsheet 2"] = list(missing_in_2)
        result["errors"] = errors

    return result

def get_used_area(sheet: Worksheet):
    """
    Given an openpyxl worksheet, returns the number of empty rows at the bottom,
    the number of empty columns on the right, as well as the last used row and column.

    This function iterates through the rows and columns from the bottom-right corner 
    and checks for non-empty values, then returns the last used row and column, 
    excluding the empty ones.

    :param sheet: The openpyxl worksheet object that contains the data.
    :type sheet: openpyxl.worksheet.worksheet.Worksheet
    :return: A dictionary containing the number of empty rows and columns at the bottom,
             and the last used row and column.
    :rtype: dict
    :raises ValueError: If the input is not a valid openpyxl worksheet.
    """

    # Validate that the input is an instance of openpyxl Worksheet
    if not isinstance(sheet, Worksheet):
        raise ValueError("The provided input is not a valid openpyxl Worksheet object.")

    # Get the last used row and column in the sheet
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Initialize counters for empty rows and columns
    empty_row_count = 0
    empty_column_count = 0

    # Count empty rows from the bottom (starting from max_row)
    for row in range(max_row, 0, -1):
        row_values = [sheet.cell(row=row, column=col).value for col in range(1, max_column + 1)]
        if all(cell is None for cell in row_values):  # If the entire row is empty
            empty_row_count += 1
        else:
            break  # Stop once a non-empty row is found

    # Count empty columns from the right (starting from max_column)
    for col in range(max_column, 0, -1):
        column_values = [sheet.cell(row=row, column=col).value for row in range(1, max_row + 1)]
        if all(cell is None for cell in column_values):  # If the entire column is empty
            empty_column_count += 1
        else:
            break  # Stop once a non-empty column is found

    # Calculate the last used row and column (excluding the empty ones)
    last_used_row = max_row - empty_row_count
    last_used_column = max_column - empty_column_count

    # Return the results as a dictionary
    return {
        "empty_rows": empty_row_count, 
        "empty_columns": empty_column_count,
        "last_used_row": last_used_row,
        "last_used_column": last_used_column,
    }


def check_sheet_structure(sheet1: Worksheet, sheet2: Worksheet, header_row_number: int = 0):
    """
    Compares the structure of two openpyxl worksheet objects to determine 
    if they have the same number of rows, columns, and column headers.

    This function validates whether the two worksheet objects are of the correct type, checks for 
    any empty sheets, compares the number of rows and columns, and ensures that the column headers 
    (both name and order) are the same in both sheets.
    It will return a detailed report indicating any discrepancies found
    between the two sheets' structures.

    Arguments:
        sheet1 (openpyxl.worksheet.worksheet.Worksheet): The first worksheet object to compare.
        sheet2 (openpyxl.worksheet.worksheet.Worksheet): The second worksheet object to compare.
        header_row_number (int, optional): The row number (1-based index) containing the
        column headers to compare. Defaults to 0, which means no header comparison will be made.

    Returns:
        dict: A dictionary containing the following structure:
            - "status" (str): Either "Error" if discrepancies were found, or "Ok"
            if the structure is identical.
            - "description" (str): A message describing the result, either listing
            discrepancies or confirming the match.
            - "errors" (dict): A dictionary with error details if discrepancies are found.
            It contains error categories (e.g., "Row/Column Count", "Empty Sheet",
            "Header Mismatch") and lists specific issues under each category.
            If no discrepancies are found, this is an empty dictionary.
    
    Example:
        sheet1 = workbook1['Sheet1']
        sheet2 = workbook2['Sheet2']
        result = check_sheet_structure(sheet1, sheet2)
        print(result)

    Notes:
        - An empty sheet is defined as one that has no rows or columns with data.
        - Column header comparison is case-sensitive and checks for exact matches
        in both name and order.
        - If `header_row_number` is set to 0, the function will skip column header comparison.
        - The function compares the maximum number of rows and columns
        (`max_row` and `max_column`) of the sheets.
    """
    errors = {}

    # Validate input types
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    # Check if both sheets are empty (either one row or one column)
    if sheet1.max_row == sheet1.max_column == sheet2.max_row == sheet2.max_column == 1:
        # Both sheets are empty, so do nothing
        pass
    else:
        # Add error for sheet1 if it's empty (either 1 row or 1 column)
        if sheet1.max_row == 1 or sheet1.max_column == 1:
            errors.setdefault("Empty Sheet", []).append(sheet1.title)

        # Add error for sheet2 if it's empty (either 1 row or 1 column)
        if sheet2.max_row == 1 or sheet2.max_column == 1:
            errors.setdefault("Empty Sheet", []).append(sheet2.title)

    # Get used area for both sheets
    shape1 = get_used_area(sheet1)
    rows1, cols1 = shape1["last_used_row"], shape1["last_used_column"]
    shape2 = get_used_area(sheet2)
    rows2, cols2 = shape2["last_used_row"], shape2["last_used_column"]

    # Check if the number of rows and columns are the same
    if (rows1, cols1) != (rows2, cols2):
        errors.setdefault("Row/Column Count", []).append(
            f"Template file has {rows1} rows and {cols1} columns, "
            f"Company file has {rows2} rows and {cols2} columns."
        )

    header1 = []
    header2 = []

    if header_row_number > 0:
        # Check if the column headers are the same (both name and order)
        header1 = [sheet1.cell(row=header_row_number, column=c).value for c in range(1, cols1 + 1)]
        header2 = [sheet2.cell(row=header_row_number, column=c).value for c in range(1, cols2 + 1)]

    if header1 != header2:
        # Find out which columns are different
        diff_headers = [
            (i + 1, h1, h2) for i, (h1, h2) in enumerate(zip(header1, header2)) if h1 != h2]
        if diff_headers:
            errors.setdefault("Header Mismatch", []).extend(
                [f"Column {i}: Template: [{h1}] != [{h2}] :Company" for i, h1, h2 in diff_headers]
            )

    # If there are errors, return "Error" status with accumulated errors
    if errors:
        return {
            "status": "Error",
            "description": "The following discrepancies were found in the sheet structure:",
            "errors": errors
        }

    # If all checks pass, return "Ok" status
    return {
        "status": "Ok",
        "description":
            f"Spreadsheets '{sheet1.title}' and '{sheet2.title}' have the same structure.",
        "errors": {}
    }

def compare_formulas(sheet1, sheet2):
    """
    Compares the formulas between two openpyxl worksheet objects.

    Arguments:
        sheet1 (openpyxl.worksheet.worksheet.Worksheet): The first worksheet to compare.
        sheet2 (openpyxl.worksheet.worksheet.Worksheet): The second worksheet to compare.

    Returns:
        dict: A dictionary with status, description, and any differences:
            - If formulas are equivalent: {
                "status": "Ok",
                "description": "All formulas are equivalent",
                "errors": {}
            }
            - If formulas differ: {
                "status": "Error",
                "description": "Found formula differences",
                "errors": {
                    "Cell_Name": ["Sheet1!A1"]
                    }
                }
    """
    # Validate input types
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    # Check if the sheets have the same number of rows and columns
    rows1, cols1 = sheet1.max_row, sheet1.max_column
    rows2, cols2 = sheet2.max_row, sheet2.max_column

    if (rows1, cols1) != (rows2, cols2):
        return {
            "status": "Error",
            "description": f"Sheets have different dimensions: '{sheet1.title}' "+\
                f"has {rows1} rows & {cols1} columns, '{sheet2.title}' has "+\
                    f"{rows2} rows & {cols2} columns.",
            "errors": {}
        }

    # Dictionary to hold differing cells, grouped by their names
    differing_cells = {}

    # Compare formulas cell by cell
    for row in range(1, rows1 + 1):
        for col in range(1, cols1 + 1):
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            # Check if both cells contain formulas (we check if cell.value starts with '=')
            if isinstance(cell1.value, str) and cell1.value.startswith('=') and \
               isinstance(cell2.value, str) and cell2.value.startswith('='):
                if cell1.value != cell2.value:
                    cell_name = f"{get_column_letter(col)}{row}"
                    # Add the differing cell to the dictionary, grouped by the cell name
                    if cell_name not in differing_cells:
                        differing_cells[cell_name] = []
                    differing_cells[cell_name].append(f"{sheet1.title}!{cell_name} "+\
                        f"({cell1.value}) != {sheet2.title}!{cell_name} ({cell2.value})")

    # If there are differences in formulas, return detailed message
    if differing_cells:
        return {
            "status": "Error",
            "description": "Found formula differences",
            "errors": differing_cells
        }

    # If all formulas are equivalent
    return {
        "status": "Ok",
        "description": "All formulas are equivalent",
        "errors": {}
    }


def check_formula_errors(sheet):
    """
    Checks for formula errors in a given openpyxl worksheet.
    
    Arguments:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to check for formula errors.
    
    Returns:
        dict: A dictionary with status, description, and any found errors in the format:
            {
                "status": "Error",
                "description": "Found errors",
                "errors": {
                    "#DIV/0!": ["Sheet1!A1"]
                }
            }
            or {"status": "Ok"} if no errors were found.
    Example:
        sheet = workbook['Sheet1']
        result = check_formula_errors(sheet)
        print(result)
    """
    # Validate input types
    if not isinstance(sheet, Worksheet):
        raise ValueError("Input must be valid openpyxl worksheet object.")

    error_details = {}

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains an error (identified by an 'e')
            if cell.data_type == 'e':
                # If the formula's output is one of the known error strings
                if isinstance(cell.value, str):
                    cell_name = f"{get_column_letter(cell.column)}{cell.row}"
                    # Group errors by type
                    if cell.value not in error_details:
                        error_details[cell.value] = []
                    error_details[cell.value].append(cell_name)

    # If no errors were found, return the status as "Ok"
    if not error_details:
        return {"status": "Ok", "description": "No errors found", "errors": {}}

    # If errors were found, return the status as "Error" with the grouped error details
    return {
        "status": "Error",
        "description": "Found errors",
        "errors": error_details
    }


# Define namedtuple for context
MissingSheetContext = namedtuple(
    'MissingSheetContext', ['Rule_Cd',
                          'Error_Category',
                          'Error_Severity_Cd']
)

def create_missing_sheet_row(sheet, context: MissingSheetContext):
    """
    Creates a dictionary representing a row for a missing sheet.
    
    Args:
        sheet (str): The name or identifier of the missing sheet.
        context (MissingSheetContext): The context containing error details like 
                                      Rule_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        dict: A dictionary containing the details of the missing sheet.
    
    Raises:
        ValueError: If 'sheet' is not a string or if 'context' does not contain
                    the required fields.
    """

    # Input validation
    if not isinstance(sheet, str) or not sheet:
        raise ValueError("The 'sheet' argument must be a non-empty string.")

    if not isinstance(context, MissingSheetContext):
        raise ValueError("The 'context' argument must be of type MissingSheetContext.")

    # Validate that context fields are not None or empty strings
    if not context.Rule_Cd or not isinstance(context.Rule_Cd, str):
        raise ValueError("Invalid 'Rule_Cd' in context: it must be a non-empty string.")

    if not context.Error_Category or not isinstance(context.Error_Category, str):
        raise ValueError("Invalid 'Error_Category' in context: it must be a non-empty string.")

    if not context.Error_Severity_Cd or not isinstance(context.Error_Severity_Cd, str):
        raise ValueError("Invalid 'Error_Severity_Cd' in context: it must be a non-empty string.")

    # Generate a unique Event_Id using uuid4
    eventid = uuid.uuid4().hex

    # Return the dictionary representing the missing sheet row
    return {
        'Event_Id': eventid,
        'Sheet_Cd': sheet,
        'Rule_Cd': context.Rule_Cd,
        'Error_Category': context.Error_Category,
        'Error_Severity_Cd': context.Error_Severity_Cd,
        'Error_Desc': "Missing Sheet"  # Description of the error
    }

def create_dataframe_missing_sheets(input_data, context: MissingSheetContext):
    """
    Creates a pandas DataFrame representing missing sheets based on
        the provided input data and context.
    
    Args:
        input_data (dict): The input data containing error details, specifically
            the list of missing sheets.
        context (MissingSheetContext): The context containing error details such as
            Rule_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        pd.DataFrame: A DataFrame containing the rows for missing sheets.
    
    Raises:
        ValueError: If 'input_data' is not a dictionary or does not contain
            the required 'errors' key.
        ValueError: If 'context' is not a valid MissingSheetContext.
    """

    # Input validation for 'input_data'
    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    # Validate that the 'context' is a valid MissingSheetContext
    if not isinstance(context, MissingSheetContext):
        raise ValueError("The 'context' argument must be of type MissingSheetContext.")

    # Extract the missing sheets list from the input_data
    missing_sheets = input_data.get('errors', {}).get('Missing In Spreadsheet 2', [])

    # Validate that 'missing_sheets' is a list
    if not isinstance(missing_sheets, list):
        missing_sheets = []  # Fallback to empty list if not a valid list

    # Create an empty list to store rows for each missing sheet
    rows = []

    # Create a row for each sheet in the missing sheets list
    for sheet in missing_sheets:
        if not isinstance(sheet, str) or not sheet:
            # pylint: disable=C0301
            raise ValueError(f"Invalid sheet name: '{sheet}'. Each sheet must be a non-empty string.")
        rows.append(create_missing_sheet_row(sheet, context))

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    return df


def find_missing_sheets(wb_template: Workbook, wb_company: Workbook):
    """
    Finds missing sheets between two provided openpyxl workbooks and returns a DataFrame 
    representing the missing sheets based on the comparison of the workbooks.
    
    Args:
        wb_template (openpyxl.workbook): The template workbook.
        wb_company (openpyxl.workbook): The company workbook.
    
    Returns:
        pd.DataFrame: A DataFrame containing rows for missing sheets.
    
    Raises:
        ValueError: If either 'wb_template' or 'wb_company' are not valid openpyxl workbooks.
    """

    # Input validation for 'wb_template' and 'wb_company'
    if not isinstance(wb_template, Workbook):
        raise ValueError("The 'wb_template' argument must be a valid openpyxl Workbook.")

    if not isinstance(wb_company, Workbook):
        raise ValueError("The 'wb_company' argument must be a valid openpyxl Workbook.")

    a = validate_tabs_between_spreadsheets(wb_template, wb_company)

    # Create the context for missing sheets
    missing_sheet_context = MissingSheetContext(
        Rule_Cd="?",
        Error_Category="Missing Sheet",
        Error_Severity_Cd="soft",
    )

    # Generate the DataFrame for missing sheets
    missing_sheets_df = create_dataframe_missing_sheets(a, missing_sheet_context)

    return missing_sheets_df

# Define the FormulaErrorSheetContext
FormulaErrorSheetContext = namedtuple(
    'FormulaErrorSheetContext', ['Rule_Cd', 'Sheet_Cd', 'Error_Category', 'Error_Severity_Cd']
)

def validate_input_data(input_data: dict, context: FormulaErrorSheetContext):
    """
    Validates the input data and context to ensure they are in the expected format.
    
    Args:
        input_data (dict): The input error data.
        context (FormulaErrorSheetContext): The context with error details.
    
    Raises:
        ValueError: If either the 'input_data' or 'context' are invalid.
    """
    # Input validation for 'input_data' and 'context'
    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    if not isinstance(context, FormulaErrorSheetContext):
        raise ValueError("The 'context' argument must be of type FormulaErrorSheetContext.")

    if any(i is None for i in context):
        raise ValueError("The 'context' values cannot be None.")

def extract_error_rows(input_data: dict):
    """
    Extracts error rows from the input data, validating the 'errors' field and its contents.
    
    Args:
        input_data (dict): The input error data.
    
    Returns:
        list: A list of tuples where each tuple contains the error type and a list of cells.
    """
    errors = input_data.get('errors', {})

    if not isinstance(errors, dict):
        raise ValueError("The 'errors' field in input_data must be a dictionary.")

    # Collect all error rows in a list
    error_rows = []
    for error_type, cells in errors.items():
        if not isinstance(cells, list):
            continue  # Skip if cells are not in list form

        error_rows.append((error_type, cells))

    return error_rows

def create_row_for_error(sheet_cd: str, error_type: str, cell:str,
                         context: FormulaErrorSheetContext):
    """
    Creates a row dictionary for a single formula error.
    
    Args:
        sheet_cd (str): The sheet code.
        error_type (str): The type of error (e.g., #DIV/0!).
        cell (str): The cell reference where the error occurred.
        context (FormulaErrorSheetContext): The context with error details.
    
    Returns:
        dict: A dictionary representing a row for the error.
    """

    return {
        'Event_Id': uuid.uuid4().hex,
        'Sheet_Cd': sheet_cd,
        'Cell_Reference': cell,
        'Rule_Cd': context.Rule_Cd,
        'Error_Category': context.Error_Category,
        'Error_Severity_Cd': context.Error_Severity_Cd,
        'Error_Desc': error_type
    }

def create_dataframe_formula_errors(input_data: dict, context: FormulaErrorSheetContext):
    """
    Creates a pandas DataFrame representing formula errors
        based on the input error data and context.
    
    Args:
        input_data (dict): A dictionary containing error details, where the keys
            are error types and the values are lists of cell references 
            affected by the errors.
        context (FormulaErrorSheetContext): A namedtuple containing error details
            like Rule_Cd, Sheet_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        pd.DataFrame: A DataFrame containing the rows for formula errors.
    
    Raises:
        ValueError: If 'input_data' is not a dictionary or if 'context' is invalid.
    """

    # Validate the input data and context
    validate_input_data(input_data, context)

    # Extract error rows from the input data
    error_rows = extract_error_rows(input_data)

    # Create the rows for the DataFrame
    rows = []
    for error_type, cells in error_rows:
        for cell in cells:
            # Create a row for each cell error
            row = create_row_for_error(context.Sheet_Cd, error_type, cell, context)
            rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    return df

def find_formula_errors(wb: Workbook):
    """
    Finds formula errors across all sheets in an Excel workbook
        and returns a consolidated DataFrame.

    Args:
        wb (Workbook): The openpyxl Workbook object representing the Excel file.

    Returns:
        pd.DataFrame: A DataFrame containing the formula errors from all sheets in the workbook.
    
    Raises:
        ValueError: If the 'wb' argument is not an instance of openpyxl Workbook.
    
    This function loops through all sheets in the provided workbook,
        runs formula checks on each sheet,
    creates a context for each sheet's errors, and generates
        a DataFrame of formula errors. The individual 
    DataFrames are then concatenated to produce one
        final DataFrame that contains all the formula errors from 
    all sheets.
    """

    # Input validation for the 'wb' argument (must be a valid openpyxl Workbook)
    if not isinstance(wb, Workbook):
        raise ValueError("The 'wb' argument must be a valid openpyxl Workbook.")

    # Initialize an empty list to store DataFrames for each sheet's formula errors
    all_formula_error_dfs = []

    # Loop through each sheet in the workbook
    for sheetname in wb.sheetnames:
        # Run formula checks for the current sheet and store the results
        formula_errors = check_formula_errors(wb[sheetname])

        # Create a context object for the current sheet with formula error details
        formula_error_sheet_context = FormulaErrorSheetContext(
            Rule_Cd="?",  # Placeholder for rule code (could be customized)
            Sheet_Cd=sheetname,
            Error_Category="Formula Error",
            Error_Severity_Cd="hard",  # Placeholder for error severity
        )

        # Create a DataFrame for the current sheet's formula errors using the helper function
        sheet_error_df = create_dataframe_formula_errors(
            formula_errors,
            formula_error_sheet_context)

        # Append the DataFrame for the current sheet to the list
        all_formula_error_dfs.append(sheet_error_df)

    # Concatenate all the DataFrames in the list into one large DataFrame
    final_formula_error_df = pd.concat(all_formula_error_dfs, ignore_index=True)

    # Return the final concatenated DataFrame containing all formula errors from all sheets
    return final_formula_error_df

# Define namedtuple for context
StructureDiscrepancyContext = namedtuple(
    'StructureDiscrepancyContext', 
    [
        'Rule_Cd',
        'Sheet_Cd',
        'Error_Category',
        'Error_Severity_Cd',
    ]
)

def create_dataframe_structure_discrepancies(
        input_data: Dict[str, Any],
        context: StructureDiscrepancyContext) -> pd.DataFrame:
    """
    Creates a DataFrame representing structure discrepancies
    between rows and columns in an Excel sheet.

    :param input_data: A dictionary containing the error data with keys
    as error types and values as the discrepancies.
    :type input_data: dict
    :param context: The context that contains information 
    like Rule Code, Sheet Code, Error Category, and Error Severity.
    :type context: StructureDiscrepancyContext

    :return: A pandas DataFrame that represents the structure discrepancies for further processing.
    :rtype: pandas.DataFrame

    :raises ValueError: If 'input_data' does not contain the expected structure
    or if the 'context' is invalid.
    :raises TypeError: If 'input_data' is not a dictionary or 'context' is
    not an instance of `StructureDiscrepancyContext`.

    This function processes the given input data (discrepancies between row/column counts) 
    and generates a DataFrame with relevant details including
    a unique event ID for each discrepancy.
    """

    # Validate input types
    if not isinstance(input_data, dict):
        raise TypeError("The 'input_data' must be a dictionary.")

    if not isinstance(context, StructureDiscrepancyContext):
        raise TypeError("The 'context' must be an instance of StructureDiscrepancyContext.")

    if 'errors' not in input_data or not isinstance(input_data['errors'], dict):
        raise ValueError("The 'input_data' must contain an 'errors' field of type dictionary.")

    # Extract context values
    rule_cd = context.Rule_Cd
    error_severity_cd = context.Error_Severity_Cd
    sheet_cd = context.Sheet_Cd
    error_category = context.Error_Category

    # Validate context values
    if not all([rule_cd, error_severity_cd, sheet_cd, error_category]):
        raise ValueError(
            "The 'context' contains missing values. Ensure all context " +\
                "attributes are properly set.")

    # Create an empty list to store rows
    rows = []

    # Extract and process structure discrepancies from the input data
    for errortype, discrepancy in input_data['errors'].items():
        # Check that the discrepancy is a list or tuple, and each element is a string
        if not isinstance(discrepancy, (list, tuple)):
            raise ValueError(
                f"The discrepancy for '{errortype}' must be a list or tuple.")

        if not all(isinstance(d, str) for d in discrepancy):
            raise ValueError(
                f"Each item in the discrepancy list for '{errortype}' must be a string.")

        # Create a row for each discrepancy (in this case, row/column count differences)
        row = {
            # Generate a unique Event ID
            'Event_Id': uuid.uuid4().hex,
            # The sheet code for the error
            'Sheet_Cd': sheet_cd,
            # Rule code (e.g., validation rule)
            'Rule_Cd': rule_cd,
            # The category of the error
            'Error_Category': error_category,
            # The severity of the error
            'Error_Severity_Cd': error_severity_cd,
            # Join the discrepancy details into a single string
            'Error_Desc': " -- ".join(discrepancy),
        }
        rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    # Return the resulting DataFrame
    return df

def find_shape_differences(wb_template: Workbook, wb_company: Workbook) -> pd.DataFrame:
    """
    Compares the sheet structures between two workbooks (template and company)
    and identifies discrepancies.

    This function checks if the sheet names exist in both workbooks, compares 
    the structures, and returns a DataFrame that highlights the discrepancies 
    found in the structures.

    :param wb_template: The template workbook to compare against.
    :type wb_template: openpyxl.Workbook
    :param wb_company: The company workbook to compare.
    :type wb_company: openpyxl.Workbook

    :return: A DataFrame containing the structure discrepancies found 
        between the two workbooks.
    :rtype: pandas.DataFrame

    :raises ValueError: If the provided workbooks are not valid or do not 
        contain any sheets.
    :raises TypeError: If the input workbooks are not instances of 
        `openpyxl.Workbook`.
    :raises KeyError: If a sheet does not exist in one of the workbooks.
    """

    # Input validation
    if not isinstance(wb_template, Workbook) or not isinstance(wb_company, Workbook):
        raise TypeError("Both inputs must be instances of openpyxl Workbook.")

    # Initialize an empty list to store individual DataFrames for discrepancies
    all_shape_error_dfs: List[pd.DataFrame] = []

    # Loop through each sheet in both workbooks and find common sheet names
    common_sheetnames = set(wb_template.sheetnames).intersection(set(wb_company.sheetnames))

    if not common_sheetnames:
        logger.warning("No common sheets found between the template and company workbooks.")

    for sheetname in common_sheetnames:
        # Create the context for the current sheet
        context = StructureDiscrepancyContext(
            Rule_Cd="?",
            Sheet_Cd=sheetname,  # Specify the sheet name with the issue
            Error_Category="Structure Discrepancy",
            Error_Severity_Cd="hard"
        )

        # only fOut_ sheets have somewhat consistent headers on row 2
        header_row_number = 2 if sheetname.startswith("fOut_") else 0

        # Check for structure discrepancies in the current sheet
        discrepancies = check_sheet_structure(
            wb_template[sheetname],
            wb_company[sheetname],
            header_row_number)

        # If discrepancies are found, create a DataFrame
        df = create_dataframe_structure_discrepancies(discrepancies, context)
        all_shape_error_dfs.append(df)

    # If no discrepancies were found, return an empty DataFrame
    if not all_shape_error_dfs:
        logger.info("No structure discrepancies were found in any sheet.")
        return pd.DataFrame()  # Return an empty DataFrame if no discrepancies

    # Concatenate all DataFrames in the list to create one big DataFrame
    final_shape_error_df = pd.concat(all_shape_error_dfs, ignore_index=True)

    # Return the final DataFrame containing all the discrepancies
    logger.info("Found %s structure discrepancies across sheets.", len(final_shape_error_df))
    return final_shape_error_df

# Define namedtuple for context
FormulaDifferencesContext = namedtuple(
    'FormulaDifferencesContext', ['Rule_Cd', 'Sheet_Cd', 'Error_Category', 'Error_Severity_Cd']
)

def create_dataframe_formula_differences(
        input_data: dict,
        context: FormulaDifferencesContext) -> pd.DataFrame:
    """
    Creates a DataFrame from input data containing formula discrepancies for a specific sheet.

    This function processes input data containing formula discrepancies
    (such as errors in formulas or missing references) and converts
    it into a pandas DataFrame. Each discrepancy will be represented as a row 
    in the DataFrame, along with associated metadata (such as the rule code,
    sheet code, error category, severity, and error description).

    :param input_data: A dictionary containing errors (keyed by cell reference)
        that occurred in the sheet. The value should be a list of error descriptions 
        for each cell.
    :type input_data: dict
    :param context: A namedtuple containing contextual information about the discrepancy, 
        including Rule_Cd, Sheet_Cd, Error_Category, and Error_Severity_Cd.
    :type context: FormulaDifferencesContext

    :return: A pandas DataFrame containing the formula discrepancies. Each row represents 
        a discrepancy with details including the event ID, sheet name, rule code, 
        cell reference, error category, severity, and description.
    :rtype: pandas.DataFrame

    :raises ValueError: If the `input_data` is not a dictionary or if the `context` is 
        not a valid `FormulaDifferencesContext`.
    :raises KeyError: If any expected key is missing in `input_data`.
    :raises TypeError: If the values in `input_data` are not lists or if any item in the 
        list is not iterable.
    """

    # Input validation
    if not isinstance(input_data, dict):
        raise ValueError("input_data must be a dictionary.")

    if not isinstance(context, FormulaDifferencesContext):
        raise ValueError("context must be an instance of FormulaDifferencesContext.")

    # Extract context values
    rule_cd = context.Rule_Cd
    error_category = context.Error_Category
    error_severity_cd = context.Error_Severity_Cd
    sheet_cd = context.Sheet_Cd

     # Validate context values
    if not all([rule_cd, error_severity_cd, sheet_cd, error_category]):
        raise ValueError(
            "The 'context' contains missing values. Ensure all context " +\
                "attributes are properly set.")

    # Create an empty list to store rows
    rows = []

    errors = input_data.get("errors", {})

    # Extract the formula discrepancies
    for cellreference, discrepancies in errors.items():
        # Create a row for each discrepancy (a list of error descriptions)
        row = {
            'Event_Id': uuid.uuid4().hex,
            'Sheet_Cd': sheet_cd,
            'Rule_Cd': rule_cd,
            'Cell_Reference': cellreference,
            'Error_Category': error_category,
            'Error_Severity_Cd': error_severity_cd,
            'Error_Desc': " -- ".join(discrepancies),  # Join the error descriptions with " -- "
        }
        rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    # Return the resulting DataFrame
    return df

def find_formula_differences(wb_template: Workbook, wb_company: Workbook) -> pd.DataFrame:
    """
    Compares the formulas between two workbooks (template and company) and identifies discrepancies.

    This function iterates through the sheets common to both workbooks and compares the formulas 
    between the sheets of the template and company workbooks. It generates a DataFrame containing 
    all formula differences (if any), including the sheet name, error category, and severity, and 
    concatenates these into a single DataFrame.

    :param wb_template: The template workbook to compare against.
    :type wb_template: openpyxl.Workbook
    :param wb_company: The company workbook to compare.
    :type wb_company: openpyxl.Workbook

    :return: A DataFrame containing all the formula differences found between the two workbooks. 
             Each row represents a formula discrepancy with details such as sheet name, 
             error category, severity, and formula description.
    :rtype: pandas.DataFrame

    :raises TypeError: If the input workbooks are not instances of `openpyxl.Workbook`.
    :raises ValueError: If either of the workbooks does not contain any sheets.
    :raises Exception: If an error occurs during the formula comparison process.
    """
    # Input validation
    if not isinstance(wb_template, Workbook) or not isinstance(wb_company, Workbook):
        raise TypeError("Both inputs must be instances of openpyxl Workbook.")

    # Initialize an empty list to store individual DataFrames
    all_formula_difference_dfs = []

    # Loop through each sheet in both workbooks and find common sheet names
    common_sheetnames = set(wb_template.sheetnames).intersection(set(wb_company.sheetnames))

    # Loop through each common sheet to compare formulas
    for sheetcd in common_sheetnames:
        # Create the context for the current sheet
        context = FormulaDifferencesContext(
            Rule_Cd="?",
            Sheet_Cd=sheetcd,  # Specify the sheet name with the issue
            Error_Category="Formula Difference",
            Error_Severity_Cd="hard"
        )

        # Compare formulas between the template and company workbooks for the current sheet
        a = compare_formulas(wb_template[sheetcd], wb_company[sheetcd])

        # Generate the DataFrame for the current sheet's formula differences
        df = create_dataframe_formula_differences(a, context)

        # Append the DataFrame for this sheet to the list
        all_formula_difference_dfs.append(df)

    # Concatenate all DataFrames in the list to create one big DataFrame
    final_formula_difference_df = pd.concat(all_formula_difference_dfs, ignore_index=True)

    # Return the final DataFrame containing all the formula differences
    return final_formula_difference_df
